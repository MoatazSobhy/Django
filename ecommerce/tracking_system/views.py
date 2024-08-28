# from django.shortcuts import render
# # from django.http import HttpResponse

# from django.http import JsonResponse
# from django.views.decorators.csrf import csrf_exempt
# import openpyxl
# import os
# import json


# # Create your views here.
# # endpoint
# def home(request):
#     # return HttpResponse("Hello")
#     return render(request=request, template_name="home.html")

# # Path to your Excel file
# FILE_PATH = r'C:\Users\Moataz Sobhy\OneDrive\Desktop\Django-Day 1\orders.xlsx'

# def get(request):
#     """Handle GET requests to read the Excel sheet."""
#     try:
#         wb = openpyxl.load_workbook(FILE_PATH)
#         sheet = wb.active
#         data = []

#         # Iterate over rows and columns in the sheet
#         for row in sheet.iter_rows(values_only=True):
#             data.append({
#                 'OrderId': row[0],
#                 'OrderName': row[1],
#                 'UserName': row[2],
#                 'Status': row[3]
#             })

#         return JsonResponse({'status': 'success', 'data': data})

#     except Exception as e:
#         return JsonResponse({'status': 'error', 'message': str(e)})

# @csrf_exempt
# def home(request):
#     # Handle GET or other requests here if needed
#     return JsonResponse({'status': 'success', 'message': 'Welcome to the home page'})

# @csrf_exempt
# def get(request):
#     # Handle GET requests here if needed
#     return JsonResponse({'status': 'success', 'message': 'GET request received'})

# @csrf_exempt
# def post(request):
#     if request.method == 'POST':
#         try:
#             # Parse the JSON data from the request body
#             data = json.loads(request.body)
            
#             # Access your data
#             order_id = data.get('OrderId')
#             order_name = data.get('OrderName')
#             user_name = data.get('UserName')
#             status = data.get('Status')
            
#             # Process the data as needed (e.g., save to database)
            
#             # Return a successful response
#             return JsonResponse({'status': 'success', 'message': 'Data received successfully'})
#         except json.JSONDecodeError:
#             return JsonResponse({'status': 'error', 'message': 'Invalid JSON'})
#     else:
#         return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
# -------------------------------------------------------------------------------------
from django.shortcuts import render
# from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import openpyxl
import json

# Path to your Excel file
FILE_PATH = r'C:\Users\Moataz Sobhy\OneDrive\Desktop\Django-Day 1\orders.xlsx'

def home(request):
    """Render the home page."""
    return render(request, template_name="home.html")

@csrf_exempt
def get(request):
    """Handle GET requests to read the Excel sheet."""
    if request.method == 'GET':
        try:
            wb = openpyxl.load_workbook(FILE_PATH)
            sheet = wb.active
            data = []

            # Iterate over rows and columns in the sheet
            for row in sheet.iter_rows(values_only=True):
                data.append({
                    'OrderId': row[0],
                    'OrderName': row[1],
                    'UserName': row[2],
                    'Status': row[3]
                })

            return JsonResponse({'status': 'success', 'data': data})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def post(request):
    """Handle POST requests to receive JSON data."""
    if request.method == 'POST':
        try:
            # Parse the JSON data from the request body
            data = json.loads(request.body)
            
            # Access your data
            order_id = data.get('OrderId')
            order_name = data.get('OrderName')
            user_name = data.get('UserName')
            status = data.get('Status')
            
            # Process the data as needed (e.g., save to database)
            
            # Return a successful response
            return JsonResponse({'status': 'success', 'message': 'Data received successfully'})
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

# -------------------------------------------------------------------------------------
# from django.shortcuts import render
# from django.http import JsonResponse
# from django.views.decorators.csrf import csrf_exempt
# import openpyxl
# import json

# # Path to your Excel file
# FILE_PATH = r'C:\Users\Moataz Sobhy\OneDrive\Desktop\Django-Day 1\orders.xlsx'

# def home(request):
#     """Render the home page."""
#     return render(request, template_name="home.html")

# @csrf_exempt
# def get(request):
#     """Handle GET requests to read the Excel sheet."""
#     if request.method == 'GET':
#         try:
#             wb = openpyxl.load_workbook(FILE_PATH)
#             sheet = wb.active
#             data = []

#             # Iterate over rows and columns in the sheet
#             for row in sheet.iter_rows(values_only=True):
#                 data.append({
#                     'OrderId': row[0],
#                     'OrderName': row[1],
#                     'UserName': row[2],
#                     'Status': row[3]
#                 })

#             return JsonResponse({'status': 'success', 'data': data})

#         except Exception as e:
#             return JsonResponse({'status': 'error', 'message': str(e)})

# @csrf_exempt
# def post(request):
#     """Handle POST requests to receive JSON data and update the Excel file."""
#     if request.method == 'POST':
#         try:
#             # Parse the JSON data from the request body
#             data = json.loads(request.body)
            
#             # Access your data
#             order_id = data.get('OrderId')
#             order_name = data.get('OrderName')
#             user_name = data.get('UserName')
#             status = data.get('Status')
            
#             # Open the Excel file
#             wb = openpyxl.load_workbook(FILE_PATH)
#             sheet = wb.active
            
#             # Append new row to the sheet
#             sheet.append([order_id, order_name, user_name, status])
            
#             # Save the workbook
#             wb.save(FILE_PATH)
            
#             # Return a successful response
#             return JsonResponse({'status': 'success', 'message': 'Data received and saved successfully'})
#         except json.JSONDecodeError:
#             return JsonResponse({'status': 'error', 'message': 'Invalid JSON'})
#         except Exception as e:
#             return JsonResponse({'status': 'error', 'message': str(e)})
#     else:
#         return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
